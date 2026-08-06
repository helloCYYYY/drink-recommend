import openpyxl, zipfile, re, os, json
from PIL import Image
from io import BytesIO

BASE = r"C:\Users\hspcadmin\WorkBuddy\饮料推荐系统"
IMG_DIR = os.path.join(BASE, "images")
os.makedirs(os.path.join(IMG_DIR, "starbucks"), exist_ok=True)
os.makedirs(os.path.join(IMG_DIR, "mstand"), exist_ok=True)

CAT_MAP = {"咖啡": "coffee", "奶茶": "milktea", "果茶": "fruittea"}

def extract(fpath, brand_key, brand_name, brand_logo):
    os.makedirs(os.path.join(IMG_DIR, brand_key), exist_ok=True)
    z = zipfile.ZipFile(fpath)
    # rels: rId -> media (attrs may appear in any order)
    relx = z.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
    rid2media = {}
    for rel in re.findall(r'<Relationship\b[^>]*/?>', relx):
        rid = re.search(r'Id="([^"]+)"', rel)
        tgt = re.search(r'Target="([^"]+)"', rel)
        if rid and tgt:
            rid2media[rid.group(1)] = tgt.group(1).lstrip("/")
    dx = z.read("xl/drawings/drawing1.xml").decode("utf-8")
    # each oneCellAnchor: row0 + rId
    anchors = re.findall(r'<oneCellAnchor>.*?</oneCellAnchor>', dx, re.S)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    brands = {}  # cat_key -> {brand_key: {name,logo,skus:[]}}
    for anc in anchors:
        m = re.search(r'<from>.*?<col>(\d+)</col>.*?<row>(\d+)</row>', anc, re.S)
        emb = re.search(r'r:embed="(rId\d+)"', anc)
        if not m or not emb:
            continue
        row0 = int(m.group(2))          # 0-based
        oprow = row0 + 1                # openpyxl row
        media = rid2media.get(emb.group(1), "")
        if media.startswith("/"):
            media = media[1:]
        # extract + re-encode image
        data = z.read(media)
        out_name = f"{oprow}.jpg"
        out_path = os.path.join(IMG_DIR, brand_key, out_name)
        try:
            im = Image.open(BytesIO(data)).convert("RGB")
            w, h = im.size
            scale = min(1.0, 600.0 / max(w, h))
            if scale < 1.0:
                im = im.resize((int(w*scale), int(h*scale)), Image.LANCZOS)
            im.save(out_path, "JPEG", quality=85)
        except Exception as e:
            # fallback: write raw bytes
            with open(out_path, "wb") as fh:
                fh.write(data)
        # read cell data
        cat = ws.cell(oprow, 1).value
        name = ws.cell(oprow, 2).value
        price = ws.cell(oprow, 4).value
        if not name:
            continue
        cat_key = CAT_MAP.get(cat)
        if not cat_key:
            print(f"  ! 未知类别 '{cat}' @row{oprow} 跳过")
            continue
        sku = {"name": str(name), "price": price if isinstance(price, (int, float)) else 0,
               "tags": [], "img": f"images/{brand_key}/{out_name}"}
        brands.setdefault(cat_key, {})
        b = brands[cat_key].setdefault(brand_key, {"name": brand_name, "logo": brand_logo, "skus": []})
        b["skus"].append(sku)
    return brands

def extract_url(fpath, brand_key, brand_name, brand_logo, price_col=4, url_col=5):
    """图床直链版：图片列直接是远程 URL，不落本地文件。

    price_col/url_col 可配：
      - 默认 4/5 → 肯悦 5 列布局(分类/名称/类型/价格/直链)
      - OT另茶 4 列布局(分类/名称/价格/直链) → price_col=3, url_col=4
    """
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    brands = {}
    skipped = 0
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        price = ws.cell(row=r, column=price_col).value
        url = ws.cell(row=r, column=url_col).value
        if not name:
            continue
        cat_key = CAT_MAP.get(cat)
        if not cat_key:
            print(f"  ! 未知类别 '{cat}' @row{r} 跳过"); skipped += 1; continue
        if not url:
            print(f"  ! 缺少图床链接 @row{r} ({name}) 跳过"); skipped += 1; continue
        sku = {"name": str(name).strip(),
               "price": price if isinstance(price, (int, float)) else 0,
               "tags": [], "img": str(url).strip()}
        brands.setdefault(cat_key, {})
        b = brands[cat_key].setdefault(brand_key, {"name": brand_name, "logo": brand_logo, "skus": []})
        b["skus"].append(sku)
    n = sum(len(b["skus"]) for cat in brands.values() for b in cat.values())
    print(f"  * {brand_key}: 读取 {n} 款 (跳过 {skipped})")
    return brands

print(">> STARBUCKS")
sb = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\STARBUCKS_饮品SKU清单_饮品居中版.xlsx",
             "starbucks", "星巴克", "🟢")
print(">> MSTAND")
ms = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\MSTAND_饮品SKU清单.xlsx",
             "mstand", "M Stand", "M")

print(">> LUCKIN")
lk = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\瑞幸咖啡_饮品SKU清单_补全版.xlsx",
             "luckin", "瑞幸", "🔵")
print(">> GUMING")
gm = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\古茗_饮品SKU清单.xlsx",
             "guming", "古茗", "🟡")
print(">> OT另茶 (图床直链)")
ot = extract_url(r"C:\Users\hspcadmin\Documents\drink\OT另茶.xlsx",
                 "otlc", "OT另茶", "🍃", price_col=3, url_col=4)
print(">> 茉莉奶白")
molly = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\茉莉奶白_饮品SKU清单.xlsx",
                "molly", "茉莉奶白", "🌿")
print(">> 裕莲茶楼")
yulian = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\裕莲茶楼_饮品SKU清单.xlsx",
                "yulian", "裕莲茶楼", "🏮")

print(">> KOI")
koi = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-07-31\ban\outputs\KOI_饮品SKU清单.xlsx",
              "koi", "KOI", "🧋")

print(">> 喜茶")
hytea = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-08-01\new-chat\outputs\喜茶饮品清单.xlsx",
                "heytea", "喜茶", "🟣")

print(">> 一点点")
yd = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-08-01\new-chat\outputs\一点点饮品清单.xlsx",
             "yidiandian", "一点点", "🔴")

print(">> O2")
o2 = extract(r"C:\Users\hspcadmin\Documents\Codex\2026-08-01\new-chat\outputs\O2饮品清单.xlsx",
             "o2", "O2", "🫧")
print(">> 肯悦咖啡 (图床直链)")
ky = extract_url(r"C:\Users\hspcadmin\Documents\drink\kenyue.xlsx",
                 "kenyue", "肯悦咖啡", "KY")
# 真实清单里没有“生椰拿铁”，但定制参数面板需要有样本：把演示 opts 挂到一杯真实拿铁上
DEMO_OPTS = {"temp": {"label": "温度", "items": ["热", "冰", "去冰"]},
             "sugar": {"label": "甜度", "items": ["全糖", "五分糖", "无糖"]},
             "topping": {"label": "加料", "items": ["珍珠", "椰果"], "multi": True}}
def attach_demo():
    target = None
    for cat in lk:
        for b in lk[cat].values():
            for s in b["skus"]:
                if s["name"] == "生椰丝绒拿铁":
                    target = s; break
            if target: break
        if target: break
    if not target:
        for b in lk.get("coffee", {}).values():
            if b["skus"]:
                target = b["skus"][0]; break
    if target:
        target["opts"] = DEMO_OPTS
        print(f"  * 定制演示 opts 挂在：{target['name']}")
attach_demo()

# ---- assemble DATA ----
DATA = {
    "coffee":   {"label": "咖啡", "emoji": "☕", "brands": {}},
    "milktea":  {"label": "奶茶", "emoji": "🥤", "brands": {}},
    "fruittea": {"label": "果茶", "emoji": "🍹", "brands": {}},
}
for cat_key in ("coffee", "milktea", "fruittea"):
    merged = {}
    for src in (sb, ms, lk, gm, ot, molly, yulian, koi, hytea, yd, o2, ky):  # 合并 十二表
        for bk, bv in src.get(cat_key, {}).items():
            merged[bk] = bv
    DATA[cat_key]["brands"] = merged

# counts
for ck in DATA:
    for bk, bv in DATA[ck]["brands"].items():
        print(f"  {ck}/{bv['name']}: {len(bv['skus'])} 款")

# ---- write data.js ----
header = '''// ============================================================
//  饮料推荐系统 · 数据文件（只改这里就能改内容）
// ============================================================
//  怎么改：
//  1. 加一杯饮品：在某品牌的 skus 数组里加一行
//       { name:'饮品名', price:价格, tags:['标签1','标签2'], img:'images/品牌/行号.jpg' }
//  2. 加一个品牌：在对应大类 brands 里加一段
//       key名: { name:'显示名', logo:'🟢', skus:[ ... ] }
//  3. 换真实图片：把图片放进 images/ 下对应品牌的文件夹，把 img 改成路径即可
//     （本项目图片来自 STARBUCKS / MSTAND 的 Excel 清单，已按行号命名）
//  4. 给某杯加定制参数（温度/甜度/加料…）：在该杯加 opts 字段，例如
//       opts: { temp:{label:'温度',items:['热','冰','去冰']},
//               sugar:{label:'甜度',items:['全糖','五分糖','无糖']},
//               topping:{label:'加料',items:['珍珠','椰果'],multi:true} }
//     不写 opts 的饮品点选后直接出结果，不弹参数面板。
//  5. 改完保存刷新页面即可生效（部署后 git push 自动更新）。
//  注意：引号用半角 " 或 '；数组最后一项不要逗号；price 用数字。
// ============================================================

const DATA = '''

with open(os.path.join(BASE, "data.js"), "w", encoding="utf-8") as f:
    f.write(header)
    json.dump(DATA, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("\n✔ data.js 已生成")
